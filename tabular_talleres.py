"""
Script para tabular los resultados del taller práctico
Fase 2 - Portafolio Docente UTN Móvil

Lee los 18 archivos .xlsm y genera un archivo consolidado resultados_talleres.xlsx
"""

import os
import datetime
import openpyxl
import pandas as pd
import numpy as np

DIRECTORIO = os.path.dirname(os.path.abspath(__file__))
ARCHIVO_SALIDA = os.path.join(DIRECTORIO, "resultados_talleres.xlsx")

ACTIVITY_ROWS = [17, 19, 21, 23, 25, 27, 29, 31]
ACTIVITY_NAMES = [
    "T1 - Ver pagos",
    "T2 - Ver control de bienes",
    "T3 - Ver Proyecto de Investigación",
    "T4 - Ver Obra de Producción Científica",
    "T5 - Comentar en Foro General",
    "T6 - Crear Evaluación Online",
    "T7 - Crear 3 preguntas",
    "T8 - Agregar Preguntas a Evaluación Online",
]


def time_to_seconds(t):
    """Convierte datetime.time a segundos totales."""
    if isinstance(t, datetime.time):
        return t.hour * 3600 + t.minute * 60 + t.second
    if isinstance(t, datetime.timedelta):
        return int(t.total_seconds())
    if isinstance(t, (int, float)):
        # Excel decimal fraction of a day
        return int(t * 86400)
    return None


def seconds_to_mmss(s):
    """Convierte segundos a formato mm:ss."""
    if s is None:
        return ""
    m, sec = divmod(int(s), 60)
    return f"{m:02d}:{sec:02d}"


def read_workshop_file(filepath):
    """Lee un archivo .xlsm y extrae los datos del taller."""
    wb = openpyxl.load_workbook(filepath, data_only=True)
    ws = wb["Datos"]

    nombre = ws["C7"].value or os.path.splitext(os.path.basename(filepath))[0]

    actividades = []
    for i, row in enumerate(ACTIVITY_ROWS):
        actividad = {
            "Participante": nombre,
            "Nro": ws[f"C{row}"].value,
            "Actividad": ACTIVITY_NAMES[i],
            "Descripcion": ws[f"D{row}"].value,
            "Inicio": ws[f"E{row}"].value,
            "Fin": ws[f"F{row}"].value,
            "Duracion_raw": ws[f"G{row}"].value,
            "Errores": ws[f"H{row}"].value,
            "Finalizo": ws[f"I{row}"].value,
            "Observaciones": ws[f"J{row}"].value,
        }

        # Calcular duración en segundos
        dur = actividad["Duracion_raw"]
        actividad["Duracion_seg"] = time_to_seconds(dur)

        actividades.append(actividad)

    wb.close()
    return nombre, actividades


def main():
    # Buscar archivos .xlsm
    archivos = sorted([
        os.path.join(DIRECTORIO, f)
        for f in os.listdir(DIRECTORIO)
        if f.endswith(".xlsm")
    ])

    print(f"Encontrados {len(archivos)} archivos .xlsm")

    all_detail = []
    resumen_rows = []

    for filepath in archivos:
        nombre_archivo = os.path.basename(filepath)
        print(f"  Leyendo: {nombre_archivo}")
        nombre, actividades = read_workshop_file(filepath)

        # Detalle
        for act in actividades:
            all_detail.append({
                "Participante": act["Participante"],
                "Nro": act["Nro"],
                "Actividad": act["Actividad"],
                "Inicio": act["Inicio"],
                "Fin": act["Fin"],
                "Duración": seconds_to_mmss(act["Duracion_seg"]),
                "Errores": act["Errores"],
                "Finalizó": act["Finalizo"],
                "Observaciones": act["Observaciones"],
            })

        # Resumen
        row = {"Participante": nombre}
        total_seg = 0
        for i, act in enumerate(actividades):
            n = i + 1
            seg = act["Duracion_seg"]
            row[f"T{n} Tiempo"] = seconds_to_mmss(seg)
            row[f"T{n} Errores"] = act["Errores"]
            row[f"T{n} Finalizó"] = act["Finalizo"]
            if seg is not None:
                total_seg += seg

        row["Tiempo Total"] = seconds_to_mmss(total_seg)
        resumen_rows.append(row)

    # DataFrames
    df_detalle = pd.DataFrame(all_detail)
    df_resumen = pd.DataFrame(resumen_rows)

    # Estadísticas por actividad
    stats_rows = []
    for i in range(8):
        n = i + 1
        nombre_act = ACTIVITY_NAMES[i]

        # Filtrar datos de esta actividad
        mask = df_detalle["Nro"] == n
        duraciones = df_detalle.loc[mask, "Duración"].apply(
            lambda x: int(x.split(":")[0]) * 60 + int(x.split(":")[1]) if x else None
        ).dropna()
        errores = df_detalle.loc[mask, "Errores"]
        finalizo = df_detalle.loc[mask, "Finalizó"]

        total_part = len(duraciones)
        pct_errores = (errores.str.upper() == "SI").sum() / len(errores) * 100 if len(errores) > 0 else 0
        pct_finalizo = (finalizo.str.upper() == "SI").sum() / len(finalizo) * 100 if len(finalizo) > 0 else 0

        stats_rows.append({
            "Actividad": nombre_act,
            "Tiempo Promedio": seconds_to_mmss(duraciones.mean()) if total_part > 0 else "",
            "Tiempo Mín": seconds_to_mmss(duraciones.min()) if total_part > 0 else "",
            "Tiempo Máx": seconds_to_mmss(duraciones.max()) if total_part > 0 else "",
            "Desv. Estándar": seconds_to_mmss(duraciones.std()) if total_part > 1 else "",
            "% con Errores": round(pct_errores, 1),
            "% Finalización Exitosa": round(pct_finalizo, 1),
            "Total Participantes": total_part,
        })

    df_stats = pd.DataFrame(stats_rows)

    # Estadísticas globales
    todos_tiempos_totales = df_resumen["Tiempo Total"].apply(
        lambda x: int(x.split(":")[0]) * 60 + int(x.split(":")[1]) if x else None
    ).dropna()
    global_row = {
        "Actividad": "TOTAL (todas las tareas)",
        "Tiempo Promedio": seconds_to_mmss(todos_tiempos_totales.mean()),
        "Tiempo Mín": seconds_to_mmss(todos_tiempos_totales.min()),
        "Tiempo Máx": seconds_to_mmss(todos_tiempos_totales.max()),
        "Desv. Estándar": seconds_to_mmss(todos_tiempos_totales.std()),
        "% con Errores": "",
        "% Finalización Exitosa": "",
        "Total Participantes": len(archivos),
    }
    df_stats = pd.concat([df_stats, pd.DataFrame([global_row])], ignore_index=True)

    # Escribir Excel
    with pd.ExcelWriter(ARCHIVO_SALIDA, engine="openpyxl") as writer:
        df_resumen.to_excel(writer, sheet_name="Resumen General", index=False)
        df_detalle.to_excel(writer, sheet_name="Detalle por Actividad", index=False)
        df_stats.to_excel(writer, sheet_name="Estadísticas", index=False)

        # Ajustar anchos de columna
        for sheet_name in writer.sheets:
            ws = writer.sheets[sheet_name]
            for col in ws.columns:
                max_len = max(len(str(cell.value or "")) for cell in col)
                col_letter = col[0].column_letter
                ws.column_dimensions[col_letter].width = min(max_len + 2, 50)

    print(f"\nArchivo generado: {ARCHIVO_SALIDA}")
    print(f"  - Resumen General: {len(df_resumen)} participantes")
    print(f"  - Detalle por Actividad: {len(df_detalle)} registros")
    print(f"  - Estadísticas: {len(df_stats)} filas")


if __name__ == "__main__":
    main()
